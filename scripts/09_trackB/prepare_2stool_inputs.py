"""
Prepare 2S-TOOL input Excel files from MLCW + GWL timeseries data.

For each station x layer combination, creates a .xlsx file with:
  - Sheet "StrainStress": Col B = displacement (m, negative = subsidence)
                          Col C = GWL depth from surface (m)
  - Sheet "InputData": hp_inicial override in cell B2 if available

Data sources (v2 — aligned timeline):
  MLCW: selectable via --mlcw-source:
        orig     → data/mlcw/group_byLayer_orig/{STATION}_orig_grouped.csv
        modeled  → data/mlcw/group_byLayer_modeled/{STATION}_modeled_grouped.csv
        reconstr → data/mlcw/group_byLayer_reconstr/{STATION}_reconst_grouped.csv
  GWL:  data/gwl/mlcw_gwl_timeseries/{MLCW_STATION}_{GWL_STATION}_{WELLCODE}.feather
        264 rows, already aligned to the MLCW timeline — no resampling needed
  Assignment: data/gwl/gwl_to_mlcw_layer_assignment_v3.csv
  Well elevations: data/gwl/well_info/gwl_allwells_flat.csv
        Use elev_leveling_m for head-to-depth conversion.
  hp_inicial overrides (optional):
        data/gwl/2stool_inputs/hp_inicial_overrides.json
        Keys: "{STATION}_{LAYER}" (e.g. "TUKU_F3"), values: float hc (m depth)

Usage:
    python scripts/09_trackB/prepare_2stool_inputs.py --stations TUKU
    python scripts/09_trackB/prepare_2stool_inputs.py --all --mlcw-source modeled
"""

import json
import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl
import logging
import sys
import argparse

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── Paths ──────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent

ASSIGNMENT_FILE = PROJECT_ROOT / "data" / "gwl" / "gwl_to_mlcw_layer_assignment_v3.csv"
ALLWELLS_FILE = PROJECT_ROOT / "data" / "gwl" / "well_info" / "gwl_allwells_flat.csv"
MLCW_GWL_TS_DIR = PROJECT_ROOT / "data" / "gwl" / "mlcw_gwl_timeseries"
OUT_DIR = PROJECT_ROOT / "data" / "gwl" / "2stool_inputs"
HP_OVERRIDES_FILE = OUT_DIR / "hp_inicial_overrides.json"

# MLCW source configuration
MLCW_SOURCE_CONFIG = {
    "orig": {
        "dir": PROJECT_ROOT / "data" / "mlcw" / "group_byLayer_orig",
        "pattern": "{station}_orig_grouped.csv",
    },
    "modeled": {
        "dir": PROJECT_ROOT / "data" / "mlcw" / "group_byLayer_modeled",
        "pattern": "{station}_modeled_grouped.csv",
    },
    "reconstr": {
        "dir": PROJECT_ROOT / "data" / "mlcw" / "group_byLayer_reconstr",
        "pattern": "{station}_reconst_grouped.csv",
    },
}


# ═══════════════════════════════════════════════════════════════════════
# Data loading utilities
# ═══════════════════════════════════════════════════════════════════════

def load_well_elevation() -> dict:
    """Load well elevation lookup from allwells_flat.csv.

    Returns
    -------
    dict: {wellcode_8digit: elev_leveling_m}
    """
    df = pd.read_csv(ALLWELLS_FILE, dtype={"wellcode": str})
    elev = {}
    for _, row in df.iterrows():
        wc = str(row["wellcode"]).strip()
        elev_m = row["elev_leveling_m"]
        if pd.notna(elev_m):
            elev[wc] = float(elev_m)
    logger.info(f"Loaded {len(elev)} well elevations")
    return elev


def load_hp_overrides() -> dict:
    """Load hp_inicial overrides from JSON if the file exists.

    Returns
    -------
    dict: {"{STATION}_{LAYER}": hc_value_float} or empty dict.
    """
    if not HP_OVERRIDES_FILE.exists():
        logger.info("No hp_inicial_overrides.json found — using data-driven hc for all layers")
        return {}
    try:
        with open(HP_OVERRIDES_FILE) as f:
            data = json.load(f)
        logger.info(f"Loaded {len(data)} hp_inicial overrides from {HP_OVERRIDES_FILE.name}")
        return data
    except Exception as exc:
        logger.warning(f"Failed to read hp_inicial_overrides.json: {exc}")
        return {}


def load_mlcw_data(station: str, source: str) -> pd.DataFrame | None:
    """Load layer-grouped MLCW data for a station from the chosen source.

    Parameters
    ----------
    station : str
        Station name.
    source : str
        One of 'orig', 'modeled', 'reconstr'.

    Returns
    -------
    pd.DataFrame with 'datetime' column + layer columns, or None if missing.
    """
    cfg = MLCW_SOURCE_CONFIG[source]
    path = cfg["dir"] / cfg["pattern"].format(station=station)
    if not path.exists():
        return None
    try:
        return pd.read_csv(path, parse_dates=["datetime"])
    except (ValueError, TypeError):
        # Fallback: datetime column may be StringDtype in some sources
        df = pd.read_csv(path)
        df["datetime"] = pd.to_datetime(df["datetime"])
        return df


def find_gwl_feather(station: str, wellcode_8: str) -> Path | None:
    """Find the GWL feather file for a (station, wellcode) pair.

    Searches MLCW_GWL_TS_DIR for files matching:
        {station}_*_{wellcode_8}.feather

    Returns the first match, or None if not found.
    """
    matches = list(MLCW_GWL_TS_DIR.glob(f"{station}_*_{wellcode_8}.feather"))
    if matches:
        return matches[0]
    return None


def load_gwl_feather(feather_path: Path) -> pd.DataFrame | None:
    """Load a GWL feather file.

    The feather has 2 columns:
      'datetime'  — already aligned to the MLCW timeline
      '{MLCW_STATION}_{GWL_STATION}_{WELLCODE}' — piezometric head in m above MSL

    Returns
    -------
    pd.DataFrame with datetime + head column, or None on error.
    """
    if not feather_path.exists():
        return None
    return pd.read_feather(feather_path)


def find_column_by_wellcode(gwl_df: pd.DataFrame, wellcode_8: str) -> str | None:
    """Find the GWL column matching a wellcode.

    New feather columns have the format {MLCW_STATION}_{GWL_STATION}_{WELLCODE}
    (e.g. 'TUKU_TUKU_09050321').  Matches by the suffix after the last underscore.
    """
    for col in gwl_df.columns:
        if col == "datetime":
            continue
        # Match by the wellcode portion after the last underscore
        parts = col.rsplit("_", 1)
        if len(parts) == 2 and parts[1] == wellcode_8:
            return col
        # Also try direct match (8-digit string)
        col_str = str(col).strip().zfill(8)
        if col_str == wellcode_8:
            return col
    return None


# ═══════════════════════════════════════════════════════════════════════
# Core: prepare one station x layer curve
# ═══════════════════════════════════════════════════════════════════════

def _get_hc_from_well_timeseries(feather_file_rel: str, wellcode_8: str,
                                  well_elev_m: float) -> float | None:
    """Derive preconsolidation head from long-term well_timeseries records.

    Reads the full historical GWL record (2000–2025) and computes hc as
    the maximum depth-to-water (elev – min_head).

    Returns None if the feather file is not found or the wellcode is missing.
    """
    if not feather_file_rel or pd.isna(feather_file_rel):
        return None

    feather_path = PROJECT_ROOT / str(feather_file_rel)
    if not feather_path.exists():
        return None

    try:
        gwl_df = pd.read_feather(feather_path)
    except Exception:
        return None

    wc_col = None
    for col in gwl_df.columns:
        if col == "datetime":
            continue
        parts = str(col).rsplit("_", 1)
        if len(parts) == 2 and parts[1] == wellcode_8:
            wc_col = col
            break
        if str(col).strip().zfill(8) == wellcode_8:
            wc_col = col
            break

    if wc_col is None:
        return None

    head = gwl_df[wc_col].dropna()
    if len(head) == 0:
        return None

    hc = well_elev_m - float(head.min())
    logger.info(
        f"  hc from well_timeseries: elev={well_elev_m:.2f} - "
        f"min_head={head.min():.2f} = {hc:.2f} m"
    )
    return hc


def prepare_one_curve(
    mlcw_df: pd.DataFrame,
    gwl_df: pd.DataFrame,
    layer: str,
    wellcode_8: str,
    well_elev_m: float,
    station: str,
    out_dir: Path,
    hp_override: float | None = None,
) -> tuple:
    """Align MLCW and GWL data for one layer and write the Excel file.

    Parameters
    ----------
    mlcw_df : pd.DataFrame
        DataFrame with 'datetime' + layer columns (cumulative mm).
    gwl_df : pd.DataFrame
        DataFrame with 'datetime' + GWL head columns (m above MSL).
    layer : str
        Layer name (column in mlcw_df).
    wellcode_8 : str
        8-digit wellcode for GWL column lookup.
    well_elev_m : float
        Well ground elevation (m) for head-to-depth conversion:
        gwl_depth_m = elev_leveling_m - piezometric_head_m
    station : str
        Station name (for output filename).
    out_dir : Path
        Output directory.
    hp_override : float or None
        If provided, write this value into the InputData sheet (cell B2)
        as the preconsolidation head override for 2S-TOOL.

    Returns
    -------
    tuple: (out_path | None, n_points, status_message)
    """
    # Validate inputs
    if layer not in mlcw_df.columns:
        return None, 0, f"Layer {layer} not in MLCW columns"

    gwcol = find_column_by_wellcode(gwl_df, wellcode_8)
    if gwcol is None:
        return None, 0, f"Wellcode {wellcode_8} not found in GWL columns {gwl_df.columns.tolist()}"

    # Align on common datetimes (inner join — should be 100% overlap since
    # both are already aligned to the same MLCW timeline)
    merged = pd.merge(
        mlcw_df[["datetime", layer]],
        gwl_df[["datetime", gwcol]],
        on="datetime",
        how="inner",
    )
    merged = merged.dropna(subset=[layer, gwcol])

    if len(merged) < 5:
        return None, len(merged), f"Too few common dates ({len(merged)})"

    # Sort chronologically
    merged = merged.sort_values("datetime")

    # Unit conversions
    displacement_m = merged[layer].values / 1000.0        # mm → m
    gwl_depth_m = well_elev_m - merged[gwcol].values      # head → depth from surface

    # Write Excel
    out_path = out_dir / f"2STOOL_{station}_{layer}.xlsx"
    wb = openpyxl.Workbook()

    # Sheet 1: StrainStress
    ws = wb.active
    ws.title = "StrainStress"
    dates_py = np.array(merged["datetime"].dt.to_pydatetime())
    for i, (dt, d, g) in enumerate(zip(dates_py, displacement_m, gwl_depth_m), start=2):
        ws.cell(row=i, column=1, value=dt)
        ws.cell(row=i, column=2, value=float(d))
        ws.cell(row=i, column=3, value=float(g))

    # Sheet 2: InputData
    ws2 = wb.create_sheet("InputData")
    if hp_override is not None:
        # Row 2, Col B = hp_inicial override
        ws2.cell(row=2, column=1, value="hp_inicial")
        ws2.cell(row=2, column=2, value=float(hp_override))

    wb.save(out_path)

    return out_path, len(merged), "OK"


# ═══════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Prepare 2S-TOOL input Excel files from aligned MLCW + GWL data."
    )
    target = parser.add_mutually_exclusive_group()
    target.add_argument(
        "--stations", type=str,
        help="Comma-separated station names (e.g. TUKU,ANHE). Default: TUKU",
        default="TUKU",
    )
    target.add_argument(
        "--all", action="store_true",
        help="Process all stations in the assignment file",
    )
    parser.add_argument(
        "--mlcw-source", type=str, choices=["orig", "modeled", "reconstr"],
        default="orig",
        help="MLCW data source (default: orig). "
             "orig=raw grouped, modeled=STL model fits, reconstr=reconstructed 5-day.",
    )
    args = parser.parse_args()

    source = args.mlcw_source
    cfg = MLCW_SOURCE_CONFIG[source]
    logger.info("MLCW source: %s (%s / %s)", source, cfg["dir"], cfg["pattern"])

    # ── Load assignment ─────────────────────────────────────────────────
    assign = pd.read_csv(ASSIGNMENT_FILE, dtype={"assigned_wellcode": str})
    all_stations = sorted(assign["station"].unique())
    logger.info(f"Assignment: {len(assign)} rows, {len(all_stations)} stations")

    if args.all:
        stations = all_stations
    else:
        stations = [s.strip() for s in args.stations.split(",")]

    unknown = [s for s in stations if s not in all_stations]
    if unknown:
        logger.error(f"Unknown stations: {unknown}. Valid: {all_stations}")
        sys.exit(1)

    # ── Load well elevation lookup ─────────────────────────────────────
    elev_lookup = load_well_elevation()

    # ── Load hp_inicial overrides ─────────────────────────────────────
    hp_overrides = load_hp_overrides()

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # ── Process station by station ─────────────────────────────────────
    ok = 0
    skipped = 0
    summary = []

    for st in stations:
        station_rows = assign[assign["station"] == st]
        logger.info(f"\n── {st} ({len(station_rows)} layers) ──")

        # -- Load MLCW data --
        mlcw_df = load_mlcw_data(st, source)
        if mlcw_df is None:
            logger.warning(f"  {st}: no MLCW file ({source} source) — skipping all layers")
            for _, row in station_rows.iterrows():
                summary.append((st, row["layer"], 0, f"MLCW {source} file missing"))
            skipped += len(station_rows)
            continue

        logger.info(f"  MLCW: {len(mlcw_df)} rows, cols={[c for c in mlcw_df.columns if c != 'datetime']}")

        # -- Process each layer --
        for _, row in station_rows.iterrows():
            ly = str(row["layer"]).strip()
            wc_8 = str(row["assigned_wellcode"]).strip().zfill(8)

            # Resolve GWL feather
            feather_path = find_gwl_feather(st, wc_8)
            if feather_path is None:
                msg = f"GWL feather not found: {st}_*_{wc_8}.feather"
                summary.append((st, ly, 0, msg))
                logger.warning(f"  {st} {ly}: SKIP — {msg}")
                skipped += 1
                continue

            gwl_df = load_gwl_feather(feather_path)
            if gwl_df is None:
                msg = f"GWL feather could not be loaded: {feather_path.name}"
                summary.append((st, ly, 0, msg))
                logger.warning(f"  {st} {ly}: SKIP — {msg}")
                skipped += 1
                continue

            # Well elevation
            well_elev = elev_lookup.get(wc_8)
            if well_elev is None:
                msg = f"well elev missing for wellcode {wc_8}"
                summary.append((st, ly, 0, msg))
                logger.warning(f"  {st} {ly}: SKIP — {msg}")
                skipped += 1
                continue

            # hp_inicial override (optional)
            hp_key = f"{st}_{ly}"
            hp_override = hp_overrides.get(hp_key)
            if hp_override is not None:
                logger.info(f"  {st} {ly}: using hp_inicial from JSON override = {hp_override:.2f} m")

            # Fall back to well_timeseries long-term record
            if hp_override is None:
                feather_file_rel = row.get("feather_file", "")
                hp_override = _get_hc_from_well_timeseries(
                    str(feather_file_rel) if pd.notna(feather_file_rel) else "",
                    wc_8, well_elev,
                )

            out_path, n_pts, status = prepare_one_curve(
                mlcw_df, gwl_df, ly, wc_8, well_elev, st, OUT_DIR,
                hp_override=hp_override,
            )

            if out_path is not None:
                summary.append((st, ly, n_pts, "OK"))
                logger.info(f"  {st} {ly}: {n_pts} pts → {out_path.name}  [feather={feather_path.name}]")
                ok += 1
            else:
                summary.append((st, ly, 0, status))
                logger.warning(f"  {st} {ly}: SKIP — {status}")
                skipped += 1

    # ── Report ──────────────────────────────────────────────────────────
    logger.info(f"\n{'=' * 60}")
    logger.info(f"Done: {ok} files written, {skipped} skipped")
    logger.info(f"Output: {OUT_DIR}")

    summary_path = OUT_DIR / "preparation_log.csv"
    pd.DataFrame(summary, columns=["station", "layer", "n_points", "status"]).to_csv(
        summary_path, index=False
    )
    logger.info(f"\nPreparation log saved: {summary_path}")

    if skipped > 0:
        logger.info("\nSkipped entries:")
        for s, ly, n, stt in summary:
            if stt != "OK":
                logger.info(f"  {s:12s} {ly:5s}: {stt}")


if __name__ == "__main__":
    main()
