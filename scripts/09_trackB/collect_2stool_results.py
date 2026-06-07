"""
Collect 2S-TOOL results across all processed output files.

Three read paths (checked in order per file):
  1. JSON fast path — reads {basename}_summary.json written by 2S-TOOL Python.
                      Preferred: NaN-safe, no pandas dependency for consumers.
  2. CSV fast path  — reads {basename}_summary.csv written by enhanced MATLAB
                      export block or 2S-TOOL Python.
  3. Excel fallback — reads the "Results" sheet of each input .xlsx file.
                      Retained for backward compatibility with runs made before
                      the JSON/CSV exports were added.

Only one read path is used per file (first successful non-empty read wins).

Outputs:
  data/gwl/2stool_outputs/2stool_results_summary.csv  — one row per (station, layer)
  data/gwl/2stool_outputs/2stool_loops_all.csv        — all per-loop data merged

Usage:
    python scripts/09_trackB/collect_2stool_results.py
"""

import json
import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl
import logging

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent

INPUT_DIR  = PROJECT_ROOT / "data" / "gwl" / "2stool_inputs"
OUTPUT_DIR = PROJECT_ROOT / "data" / "gwl" / "2stool_outputs"


# ── Filename helpers ────────────────────────────────────────────────────

def parse_filename(path: Path) -> dict | None:
    """Return {'station': ..., 'layer': ...} from 2STOOL_{STATION}_{LAYER}.xlsx."""
    parts = path.stem.split("_", 2)
    if len(parts) == 3 and parts[0] == "2STOOL":
        return {"station": parts[1], "layer": parts[2]}
    return None


# ── JSON fast path (preferred) ────────────────────────────────────────────

def read_summary_json(input_xlsx: Path) -> dict | None:
    """Read the _summary.json produced by 2S-TOOL Python.

    Checks OUTPUT_DIR/{stem}/ folder (per-file subfolder).
    Returns None if the JSON file does not exist.
    """
    json_path = OUTPUT_DIR / input_xlsx.stem / (input_xlsx.stem + "_summary.json")
    if not json_path.exists():
        return None
    try:
        with open(json_path) as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError):
        return None
    if not data:
        return {"error": "empty summary JSON"}
    return {
        "skv":                    _float(data.get("skv")),
        "ske_max":                _float(data.get("ske_max")),
        "ske_mean":               _float(data.get("ske_mean")),
        "ske_min":                _float(data.get("ske_min")),
        "ske_weighted":           _float(data.get("ske_weighted")),
        "ske_std":                _float(data.get("ske_std")),
        "n_loops":                _int(data.get("n_loops_total")),
        "n_accepted_loops":       _int(data.get("n_loops_accepted")),
        "y_interval":             _float(data.get("y_interval")),
        "x_interval":             _float(data.get("x_interval")),
        "preconsolidation_depth_hc": _float(data.get("hc")),
        "pct_amplitude":          _float(data.get("pct_amplitude")),
        "source": "json",
    }


def read_loops_json(input_xlsx: Path) -> pd.DataFrame | None:
    """Read the _loops.json produced by 2S-TOOL Python."""
    json_path = OUTPUT_DIR / input_xlsx.stem / (input_xlsx.stem + "_loops.json")
    if not json_path.exists():
        return None
    try:
        with open(json_path) as f:
            records = json.load(f)
    except (json.JSONDecodeError, OSError):
        return None
    if not records:
        return None
    return pd.DataFrame(records)


# ── CSV fast path ───────────────────────────────────────────────────────

def read_summary_csv(input_xlsx: Path) -> dict | None:
    """Read the _summary.csv produced by the enhanced MATLAB export block.

    Checks OUTPUT_DIR/{stem}/ folder first (per-file subfolder), then
    OUTPUT_DIR/ flat (legacy).  Returns None if neither exists.
    """
    for csv_path in [
        OUTPUT_DIR / input_xlsx.stem / (input_xlsx.stem + "_summary.csv"),
        OUTPUT_DIR / (input_xlsx.stem + "_summary.csv"),
    ]:
        if csv_path.exists():
            break
    else:
        return None
    try:
        df = pd.read_csv(csv_path)
        if df.empty:
            return {"error": "empty summary CSV"}
        row = df.iloc[0]
        return {
            "skv":                    _float(row.get("skv")),
            "ske_max":                _float(row.get("ske_max")),
            "ske_mean":               _float(row.get("ske_mean")),
            "ske_min":                _float(row.get("ske_min")),
            "ske_weighted":           _float(row.get("ske_weighted")),
            "ske_std":                _float(row.get("ske_std")),
            "n_loops":                _int(row.get("n_loops_total")),
            "n_accepted_loops":       _int(row.get("n_loops_accepted")),
            "y_interval":             _float(row.get("y_interval")),
            "x_interval":             _float(row.get("x_interval")),
            "preconsolidation_depth_hc": _float(row.get("hc")),
            "pct_amplitude":          _float(row.get("pct_amplitude")),
            "source": "csv",
        }
    except Exception as exc:
        return {"error": f"CSV parse error: {exc}"}


def read_loops_csv(input_xlsx: Path) -> pd.DataFrame | None:
    """Read the _loops.csv produced by the enhanced MATLAB export block."""
    for csv_path in [
        OUTPUT_DIR / input_xlsx.stem / (input_xlsx.stem + "_loops.csv"),
        OUTPUT_DIR / (input_xlsx.stem + "_loops.csv"),
    ]:
        if csv_path.exists():
            try:
                return pd.read_csv(csv_path)
            except Exception:
                return None
    return None


# ── Excel fallback ──────────────────────────────────────────────────────

def read_results_excel(input_xlsx: Path) -> dict | None:
    """Read 2S-TOOL Results sheet.

    Checks OUTPUT_DIR first (copy written there by updated MATLAB), then
    falls back to INPUT_DIR (older runs where output landed in the same dir).
    Returns None if the file has not been processed by MATLAB yet.
    Returns {'error': ...} on read failure.
    """
    # Check per-file subfolder first, then flat OUTPUT_DIR, then INPUT_DIR
    for filepath in [
        OUTPUT_DIR / input_xlsx.name,                       # legacy flat
        OUTPUT_DIR / input_xlsx.stem / input_xlsx.name,      # per-file subfolder
        input_xlsx,                                          # original input
    ]:
        if filepath.exists():
            break
    else:
        return None
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception as exc:
        return {"error": f"Cannot open: {exc}"}

    if "Results" not in wb.sheetnames:
        wb.close()
        return None

    ws = wb["Results"]

    params = {}
    for r, name in [(3, "y_interval"), (4, "x_interval"),
                    (5, "preconsolidation_depth_hc"), (6, "pct_amplitude")]:
        val = ws.cell(row=r, column=6).value
        params[name] = _float(val)

    agg = {}
    for r, name in [(10, "skv"), (11, "ske_max"),
                    (12, "ske_mean"), (13, "ske_min"), (14, "ske_weighted")]:
        val = ws.cell(row=r, column=6).value
        agg[name] = _float(val)

    n_accepted = 0
    n_loops = 0
    r = 3
    while True:
        ske_val = ws.cell(row=r, column=1).value
        if ske_val is None:
            break
        n_loops += 1
        if ws.cell(row=r, column=3).value == 1:
            n_accepted += 1
        r += 1

    wb.close()

    return {
        **params,
        **agg,
        "ske_std": None,
        "n_loops": n_loops,
        "n_accepted_loops": n_accepted,
        "source": "excel",
    }


# ── Type coercions ──────────────────────────────────────────────────────

def _float(v) -> float | None:
    try:
        return float(v) if v is not None and not (isinstance(v, float) and np.isnan(v)) else None
    except (TypeError, ValueError):
        return None


def _int(v) -> int | None:
    try:
        return int(v) if v is not None else None
    except (TypeError, ValueError):
        return None


# ── Main ────────────────────────────────────────────────────────────────

def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    files = sorted(INPUT_DIR.glob("2STOOL_*.xlsx"))
    logger.info(f"Found {len(files)} input files in {INPUT_DIR}")

    results = []
    all_loops = []
    n_json = n_csv = n_excel = n_not_processed = n_errors = 0

    for fp in files:
        meta = parse_filename(fp)
        if meta is None:
            continue

        # Try JSON fast path first (preferred — NaN-safe, pandas-optional)
        data = read_summary_json(fp)
        if data is not None and "error" not in data:
            n_json += 1
        else:
            if data and "error" in data:
                logger.warning(f"  {meta['station']} {meta['layer']}: {data['error']} — trying CSV")
            # Try CSV fast path
            data = read_summary_csv(fp)
            if data is not None and "error" not in data:
                n_csv += 1
            else:
                if data and "error" in data:
                    logger.warning(f"  {meta['station']} {meta['layer']}: {data['error']} — trying Excel")
                data = read_results_excel(fp)
                if data is None:
                    n_not_processed += 1
                    continue
                if "error" in data:
                    logger.warning(f"  {meta['station']} {meta['layer']}: {data['error']}")
                    n_errors += 1
                    continue
                n_excel += 1

        results.append({
            "station": meta["station"],
            "layer":   meta["layer"],
            "file":    fp.name,
            **{k: v for k, v in data.items() if k != "source"},
            "source":  data["source"],
        })

        # Collect per-loop data (try JSON first, then CSV)
        loops_df = read_loops_json(fp)
        if loops_df is None:
            loops_df = read_loops_csv(fp)
        if loops_df is not None:
            loops_df.insert(0, "station", meta["station"])
            loops_df.insert(1, "layer",   meta["layer"])
            all_loops.append(loops_df)

    logger.info(
        f"\nRead via JSON: {n_json}  |  CSV: {n_csv}  |  "
        f"Excel fallback: {n_excel}  |  "
        f"Not yet run: {n_not_processed}  |  Errors: {n_errors}"
    )

    if not results:
        logger.info("No files have been processed by 2S-TOOL yet.")
        logger.info("Run 2S-TOOL (MATLAB) first, then re-run this script.")
        return

    # ── Write summary ─────────────────────────────────────────────────
    df = pd.DataFrame(results)
    summary_path = OUTPUT_DIR / "2stool_results_summary.csv"
    df.to_csv(summary_path, index=False)
    logger.info(f"\nSummary: {summary_path}")
    logger.info(f"  {len(df)} layers, {df['station'].nunique()} stations")

    # ── Write merged loops ────────────────────────────────────────────
    if all_loops:
        loops_merged = pd.concat(all_loops, ignore_index=True)
        loops_path = OUTPUT_DIR / "2stool_loops_all.csv"
        loops_merged.to_csv(loops_path, index=False)
        logger.info(f"  Loops: {loops_path}  ({len(loops_merged)} rows from {len(all_loops)} files)")

    # ── Quick diagnostics ─────────────────────────────────────────────
    skv_valid = df["skv"].dropna()
    ske_valid = df["ske_mean"].dropna()
    ske_w_valid = df["ske_weighted"].dropna()

    if not skv_valid.empty:
        logger.info(f"\nS_kv range:    {skv_valid.min():.4e} to {skv_valid.max():.4e}")
    if not ske_valid.empty:
        logger.info(f"S_ke range:    {ske_valid.min():.4e} to {ske_valid.max():.4e}")
    if not ske_w_valid.empty:
        logger.info(f"S_ke weighted: {ske_w_valid.min():.4e} to {ske_w_valid.max():.4e}")

    accepted_col = df["n_accepted_loops"].dropna()
    if not accepted_col.empty:
        logger.info(f"Accepted loops: {int(accepted_col.min())} to {int(accepted_col.max())} per layer")

    zero_accepted = df[df["n_accepted_loops"] == 0]
    if not zero_accepted.empty:
        logger.warning(f"\nLayers with 0 accepted loops ({len(zero_accepted)}):")
        for _, r in zero_accepted.iterrows():
            logger.warning(f"  {r['station']} {r['layer']} — consider lowering y_interval or pct_amplitude")


if __name__ == "__main__":
    main()
